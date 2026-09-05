import os
os.environ["OPENBLAS_NUM_THREADS"] = "1"

import sqlite3
from datetime import datetime
import io
import pandas as pd
from PIL import Image
import streamlit as st

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. Configuración de página
logo_icon = None
for posible_logo in ["logo.jpg", "logo.png", "logo.jpeg", "21237.jpg"]:
    if os.path.exists(posible_logo):
        try:
            logo_icon = Image.open(posible_logo)
            break
        except Exception:
            pass

st.set_page_config(
    page_title="Minici Store",
    page_icon=logo_icon if logo_icon else "🛍️",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# 2. Base de datos SQLite
conn = sqlite3.connect("minici_store.db", check_same_thread=False)
c = conn.cursor()

c.execute("""CREATE TABLE IF NOT EXISTS clientes (
               id_cliente TEXT PRIMARY KEY,
               nombre TEXT NOT NULL,
               telefono TEXT,
               correo TEXT)""")

c.execute("""CREATE TABLE IF NOT EXISTS productos (
               id INTEGER PRIMARY KEY AUTOINCREMENT,
               id_cliente TEXT,
               tienda TEXT,
               descripcion TEXT,
               precio REAL DEFAULT 0.0,
               moneda TEXT DEFAULT 'CRC',
               cantidad INTEGER DEFAULT 1,
               estado TEXT,
               id_caja TEXT,
               observaciones TEXT,
               foto_path TEXT,
               FOREIGN KEY(id_cliente) REFERENCES clientes(id_cliente))""")

columnas_necesarias = [
    ("categoria", "TEXT"),
    ("precio", "REAL DEFAULT 0.0"),
    ("moneda", "TEXT DEFAULT 'CRC'"),
    ("cantidad", "INTEGER DEFAULT 1"),
    ("id_caja", "TEXT"),
    ("observaciones", "TEXT"),
    ("foto_path", "TEXT")
]

for col_name, col_type in columnas_necesarias:
    try:
        c.execute(f"ALTER TABLE productos ADD COLUMN {col_name} {col_type}")
        conn.commit()
    except sqlite3.OperationalError:
        pass

c.execute("""CREATE TABLE IF NOT EXISTS abonos (
               id INTEGER PRIMARY KEY AUTOINCREMENT,
               id_cliente TEXT,
               monto_crc REAL,
               fecha TEXT)""")

c.execute("""CREATE TABLE IF NOT EXISTS gastos (
               id INTEGER PRIMARY KEY AUTOINCREMENT,
               concepto TEXT NOT NULL,
               categoria TEXT,
               monto_crc REAL NOT NULL,
               fecha TEXT NOT NULL,
               observaciones TEXT)""")

c.execute("""CREATE TABLE IF NOT EXISTS notificaciones (
               id INTEGER PRIMARY KEY AUTOINCREMENT,
               id_cliente TEXT,
               titulo TEXT,
               mensaje TEXT,
               leida INTEGER DEFAULT 0,
               fecha TEXT)""")
conn.commit()

if not os.path.exists("fotos_productos"):
    os.makedirs("fotos_productos")

# 3. Estilos CSS Globales
st.markdown(
    """
<style>
   .stApp {
       background-color: #fdf2f8 !important;
       font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
       color: #0f172a !important;
   }
   
   #MainMenu, footer, header {visibility: hidden;}

   label, p, span, div, h1, h2, h3, h4, h5, h6 {
       color: #0f172a;
   }

   /* METRICAS */
   div[data-testid="stMetric"] {
       background-color: #ffffff !important;
       padding: 12px 16px !important;
       border-radius: 12px !important;
       border: 1px solid #fbcfe8 !important;
       box-shadow: 0 2px 6px rgba(0, 0, 0, 0.04) !important;
   }
   div[data-testid="stMetricLabel"] > div,
   div[data-testid="stMetricLabel"] label,
   div[data-testid="stMetricLabel"] p {
       color: #9d174d !important;
       font-weight: 700 !important;
       font-size: 12px !important;
       text-transform: uppercase !important;
   }
   div[data-testid="stMetricValue"] > div {
       color: #831843 !important;
       font-weight: 800 !important;
   }

   /* Banner superior */
   .top-banner {
       background: linear-gradient(135deg, #f472b6 0%, #db2777 100%);
       padding: 18px;
       border-radius: 14px;
       color: #ffffff !important;
       margin-bottom: 20px;
       box-shadow: 0 4px 12px rgba(219, 39, 119, 0.15);
   }
   .top-banner * {
       color: #ffffff !important;
   }

   /* Tarjetas blancas */
   .form-card {
       background: #ffffff !important;
       padding: 22px;
       border-radius: 14px;
       box-shadow: 0 2px 10px rgba(0, 0, 0, 0.05);
       margin-bottom: 20px;
       border: 1px solid #fbcfe8;
   }

   .section-title {
       font-size: 13px;
       font-weight: 700;
       color: #be185d !important;
       margin-bottom: 6px;
       margin-top: 10px;
       text-transform: uppercase;
   }

   div[data-baseweb="select"],
   div[data-baseweb="select"] *,
   div[data-baseweb="input"],
   div[data-baseweb="input"] *,
   div[data-baseweb="base-input"],
   div[data-baseweb="base-input"] *,
   .stSelectbox > div > div,
   .stTextInput > div > div,
   .stNumberInput > div > div {
       background-color: #ffffff !important;
       color: #0f172a !important;
       fill: #0f172a !important;
   }

   input, textarea, [role="option"], [role="combobox"] {
       color: #0f172a !important;
       -webkit-text-fill-color: #0f172a !important;
       background-color: #ffffff !important;
   }

   div[data-baseweb="popover"],
   div[data-baseweb="popover"] * {
       background-color: #ffffff !important;
       color: #0f172a !important;
   }

   .stNumberInput button {
       background-color: #fbcfe8 !important;
       color: #be185d !important;
       border: none !important;
   }

   /* Botones principales */
   div.stButton > button:first-child, div.stDownloadButton > button:first-child {
       background-color: #ec4899 !important;
       color: #ffffff !important;
       border-radius: 10px;
       font-weight: 700;
       border: none;
       padding: 10px 18px;
       width: 100%;
       font-size: 15px;
   }
   div.stButton > button:first-child *, div.stDownloadButton > button:first-child * {
       color: #ffffff !important;
   }
   div.stButton > button:first-child:hover, div.stDownloadButton > button:first-child:hover {
       background-color: #db2777 !important;
   }

   /* Radio Buttons */
   div[data-testid="stRadio"] > div {
       flex-direction: row !important;
       gap: 8px !important;
       flex-wrap: wrap !important;
   }
   
   div[data-testid="stRadio"] label {
       background-color: #ffffff !important;
       border: 2px solid #f472b6 !important;
       padding: 8px 14px !important;
       border-radius: 10px !important;
       font-weight: 700 !important;
       cursor: pointer !important;
   }

   div[data-testid="stRadio"] label p,
   div[data-testid="stRadio"] label span,
   div[data-testid="stRadio"] label div {
       color: #db2777 !important;
       -webkit-text-fill-color: #db2777 !important;
   }

   div[data-testid="stRadio"] label:has(input:checked) {
       background-color: #db2777 !important;
       border-color: #db2777 !important;
   }

   div[data-testid="stRadio"] label:has(input:checked) p,
   div[data-testid="stRadio"] label:has(input:checked) span,
   div[data-testid="stRadio"] label:has(input:checked) div {
       color: #ffffff !important;
       -webkit-text-fill-color: #ffffff !important;
   }

   div[data-testid="stRadio"] input[type="radio"] {
       display: none !important;
   }
</style>
""",
    unsafe_allow_html=True,
)

# 4. Control de Sesión
if "user_role" not in st.session_state:
    st.session_state.user_role = None
if "current_client" not in st.session_state:
    st.session_state.current_client = None

# --- INICIO DE SESIÓN ---
if st.session_state.user_role is None:
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        logo_path = None
        for posible_nombre in ["logo.jpg", "logo.png", "logo.jpeg", "21237.jpg"]:
            if os.path.exists(posible_nombre):
                logo_path = posible_nombre
                break

        if logo_path:
            st.image(logo_path, use_container_width=True)
        else:
            st.markdown(
                "<h1 style='text-align: center; color: #db2777;'>🛍️ Minici Store</h1>",
                unsafe_allow_html=True,
            )

    st.markdown(
        """
        <div style="text-align: center; margin-bottom: 15px;">
            <h2 style="color: #db2777; font-size: 22px;">¡Bienvenidos a Minici Store!</h2>
            <p style="color: #64748b; font-size: 14px;">Ingresa tu código de acceso para continuar.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<div class="form-card">', unsafe_allow_html=True)
    codigo_ingresado = st.text_input("Código de acceso", placeholder="Ej. MIN-0001")

    st.write("")
    if st.button("🚀 Ingresar al Sistema"):
        codigo_limpio = codigo_ingresado.strip().upper()
        
        if codigo_limpio == "KENDRA5412":
            st.session_state.user_role = "admin"
            st.rerun()
        else:
            c.execute("SELECT id_cliente FROM clientes WHERE UPPER(id_cliente) = ?", (codigo_limpio,))
            res = c.fetchone()
            if res:
                st.session_state.user_role = "client"
                st.session_state.current_client = res[0]
                st.rerun()
            else:
                st.error("❌ Código incorrecto o no registrado.")
    st.markdown("</div>", unsafe_allow_html=True)

# --- PANEL ADMINISTRADOR ---
elif st.session_state.user_role == "admin":
    col_a, col_b = st.columns([3, 1])
    with col_a:
        st.markdown(
            "<h3 style='color: #db2777;'>⚙️ Panel Administrador</h3>", unsafe_allow_html=True
        )
    with col_b:
        if st.button("🚪 Salir"):
            st.session_state.user_role = None
            st.rerun()

    menu_admin = st.radio(
        "Acción Admin",
        [
            "📸 Registrar Compra",
            "👩 Registrar Clienta",
            "📦 Control de Cajas",
            "💰 Registrar Abonos",
            "💸 Gastos Operativos",
            "📊 Finanzas y Reportes"
        ],
        horizontal=True,
        label_visibility="collapsed",
    )
    st.write("")

    if menu_admin == "📸 Registrar Compra":
        st.markdown(
            """
            <div class="top-banner">
                <div style="font-size: 18px; font-weight: 700;">Registrar nueva compra</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        st.markdown('<div class="form-card">', unsafe_allow_html=True)
        clientes_df = pd.read_sql(
            "SELECT id_cliente || ' — ' || nombre AS display, id_cliente FROM clientes", conn
        )

        if clientes_df.empty:
            st.warning("No hay clientas registradas.")
        else:
            st.markdown('<p class="section-title">1. Seleccionar clienta</p>', unsafe_allow_html=True)
            cli_selected = st.selectbox(
                "Clienta", clientes_df["display"], label_visibility="collapsed"
            )
            id_cliente = clientes_df[clientes_df["display"] == cli_selected]["id_cliente"].values[0]

            col_tienda, col_cat = st.columns(2)
            with col_tienda:
                st.markdown('<p class="section-title">2. Tienda</p>', unsafe_allow_html=True)
                tienda_sel = st.selectbox(
                    "Tienda",
                    [
                        "Zara",
                        "Guess",
                        "Adidas",
                        "Shein",
                        "Amazon",
                        "Nike",
                        "Victoria's Secret",
                        "Otra",
                    ],
                    label_visibility="collapsed",
                )
                if tienda_sel == "Otra":
                    tienda = st.text_input(
                        "Escribe el nombre de la tienda", placeholder="Escribe la tienda..."
                    )
                else:
                    tienda = tienda_sel

            with col_cat:
                st.markdown('<p class="section-title">3. Categoría</p>', unsafe_allow_html=True)
                cat_sel = st.selectbox(
                    "Categoría",
                    [
                        "Vestido",
                        "Bolso",
                        "Tenis",
                        "Blusa",
                        "Cosméticos",
                        "Accesorios",
                        "Otra",
                    ],
                    label_visibility="collapsed",
                )
                if cat_sel == "Otra":
                    categoria = st.text_input(
                        "Escribe el nombre de la categoría", placeholder="Escribe la categoría..."
                    )
                else:
                    categoria = cat_sel

            st.markdown('<p class="section-title">4. Producto</p>', unsafe_allow_html=True)
            producto = st.text_input(
                "Producto", placeholder="Ej. Vestido estampado", label_visibility="collapsed"
            )

            col_c, col_d = st.columns(2)
            with col_c:
                st.markdown(
                    '<p class="section-title">5. Precio (₡ CRC)</p>', unsafe_allow_html=True
                )
                precio = st.number_input(
                    "Precio",
                    min_value=0.0,
                    value=15000.0,
                    step=500.0,
                    label_visibility="collapsed",
                )
            with col_d:
                st.markdown('<p class="section-title">6. Cantidad</p>', unsafe_allow_html=True)
                cantidad = st.number_input(
                    "Cantidad", min_value=1, value=1, step=1, label_visibility="collapsed"
                )

            col_est, col_caja = st.columns(2)
            with col_est:
                st.markdown(
                    '<p class="section-title">7. Estado del producto</p>', unsafe_allow_html=True
                )
                estado = st.selectbox(
                    "Estado",
                    [
                        "🇺🇸 Comprado en USA",
                        "📦 En tránsito",
                        "🇨🇷 Recibido en CR",
                        "✅ Entregado",
                    ],
                    label_visibility="collapsed",
                )
            with col_caja:
                st.markdown('<p class="section-title">8. Asignar a caja</p>', unsafe_allow_html=True)
                caja = st.selectbox(
                    "Caja",
                    ["Seleccionar caja", "Caja C01", "Caja C02", "Caja C03", "Caja C04"],
                    label_visibility="collapsed",
                )

            st.markdown(
                '<p class="section-title">9. Fotos del producto</p>', unsafe_allow_html=True
            )
            metodo_foto = st.radio("Cargar foto desde:", ["Subir archivo", "Usar cámara"], horizontal=True)
            foto_file = None
            if metodo_foto == "Subir archivo":
                foto_file = st.file_uploader("Subir foto", type=["jpg", "png", "jpeg"], label_visibility="collapsed")
            else:
                foto_file = st.camera_input("Tomar foto")

            st.markdown(
                '<p class="section-title">10. Observaciones</p>', unsafe_allow_html=True
            )
            observaciones = st.text_area(
                "Observaciones", height=70, label_visibility="collapsed"
            )

            st.write("")
            if st.button("💾 Guardar compra"):
                if not producto or not tienda or not categoria:
                    st.error("Debes completar el producto, la tienda y la categoría.")
                else:
                    foto_filename = ""
                    if foto_file:
                        foto_filename = (
                            f"{id_cliente}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
                        )
                        filepath = os.path.join("fotos_productos", foto_filename)
                        with open(filepath, "wb") as f:
                            f.write(foto_file.getbuffer())

                    c.execute(
                        """INSERT INTO productos (id_cliente, tienda, categoria, descripcion, precio, moneda, cantidad, estado, id_caja, observaciones, foto_path)
                                     VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (
                            id_cliente,
                            tienda,
                            categoria,
                            producto,
                            precio,
                            "CRC",
                            cantidad,
                            estado,
                            caja,
                            observaciones,
                            foto_filename,
                        ),
                    )

                    c.execute(
                        """INSERT INTO notificaciones (id_cliente, titulo, mensaje, fecha)
                                     VALUES (?, ?, ?, ?)""",
                        (
                            id_cliente,
                            "Nuevo pedido registrado",
                            f"Se agregó '{producto}' ({tienda}) a tus compras por ₡{precio:,.0f}.",
                            datetime.now().strftime("%Y-%m-%d %H:%M"),
                        ),
                    )
                    conn.commit()
                    st.success(f"¡Compra de '{producto}' registrada exitosamente!")

        st.markdown("</div>", unsafe_allow_html=True)

    elif menu_admin == "👩 Registrar Clienta":
        st.markdown(
            """
            <div class="top-banner">
                <div style="font-size: 18px; font-weight: 700;">Registrar Nueva Clienta</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        st.markdown('<div class="form-card">', unsafe_allow_html=True)
        c.execute("SELECT id_cliente FROM clientes")
        rows = c.fetchall()
        
        numeros = [
            int(r[0].replace("MIN-", ""))
            for r in rows
            if r[0] and r[0].startswith("MIN-") and r[0].replace("MIN-", "").isdigit()
        ]
        siguiente_num = max(numeros) + 1 if numeros else 1
        nuevo_id = f"MIN-{siguiente_num:04d}"

        st.markdown(
            f"🏷️ <span style='font-size:16px; font-weight:bold; color:#be185d;'>Código asignado: {nuevo_id}</span>",
            unsafe_allow_html=True,
        )
        st.write("")

        nombre = st.text_input("Nombre de la Clienta", placeholder="Ej. Maria Lopez")
        tel = st.text_input("Teléfono / WhatsApp", placeholder="Ej. 88888888")
        correo = st.text_input("Correo Electrónico", placeholder="Ej. correo@ejemplo.com")

        st.write("")
        if st.button("Guardar Clienta"):
            if nombre:
                c.execute(
                    "INSERT OR REPLACE INTO clientes (id_cliente, nombre, telefono, correo) VALUES (?, ?, ?, ?)",
                    (nuevo_id, nombre, tel, correo),
                )
                conn.commit()
                st.success(f"¡Clienta {nombre} guardada exitosamente ({nuevo_id})!")
            else:
                st.error("Debes ingresar el nombre de la clienta.")
        st.markdown("</div>", unsafe_allow_html=True)

    elif menu_admin == "📦 Control de Cajas":
        st.markdown('<div class="form-card">', unsafe_allow_html=True)
        cajas_df = pd.read_sql(
            "SELECT DISTINCT id_caja FROM productos WHERE id_caja IS NOT NULL AND id_caja != 'Seleccionar caja'", conn
        )
        if cajas_df.empty:
            st.info("No hay cajas registradas con productos.")
        else:
            caja_sel = st.selectbox("Seleccionar Caja", cajas_df["id_caja"])
            items_caja = pd.read_sql(
                "SELECT p.id_cliente, c.nombre, p.descripcion, p.precio as precio_crc, p.estado FROM productos p JOIN clientes c ON p.id_cliente = c.id_cliente WHERE p.id_caja = ?",
                conn,
                params=(caja_sel,),
            )
            st.dataframe(items_caja, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    # --- CAMBIO 1: REGISTRO Y CONSULTA DE ABONOS POR CLIENTE ---
    elif menu_admin == "💰 Registrar Abonos":
        st.markdown(
            """
            <div class="top-banner">
                <div style="font-size: 18px; font-weight: 700;">Gestión de Abonos</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        st.markdown('<div class="form-card">', unsafe_allow_html=True)
        cli_list = pd.read_sql(
            "SELECT id_cliente || ' - ' || nombre AS disp, id_cliente FROM clientes", conn
        )
        if cli_list.empty:
            st.warning("No hay clientas registradas.")
        else:
            sel = st.selectbox("Seleccionar Clienta para Abonar", cli_list["disp"])
            id_c = sel.split(" - ")[0]
            monto = st.number_input("Monto Abonado (₡ CRC)", min_value=0.0, step=1000.0)
            st.write("")
            if st.button("Guardar Abono"):
                c.execute(
                    "INSERT INTO abonos (id_cliente, monto_crc, fecha) VALUES (?, ?, ?)",
                    (id_c, monto, datetime.now().strftime("%Y-%m-%d %H:%M")),
                )
                conn.commit()
                st.success("Abono registrado con éxito.")

        st.markdown("</div>", unsafe_allow_html=True)

        # Historial visual de Abonos por Clienta
        st.markdown('<div class="form-card">', unsafe_allow_html=True)
        st.markdown("<h4 style='color:#be185d;'>📋 Historial de Abonos Recibidos</h4>", unsafe_allow_html=True)
        
        abonos_detalle = pd.read_sql("""
            SELECT 
                a.fecha as Fecha, 
                c.id_cliente as Código, 
                c.nombre as Clienta, 
                a.monto_crc as 'Monto Abonado (₡)' 
            FROM abonos a 
            JOIN clientes c ON a.id_cliente = c.id_cliente 
            ORDER BY a.id DESC
        """, conn)

        if not abonos_detalle.empty:
            st.dataframe(abonos_detalle, use_container_width=True)
        else:
            st.info("No hay abonos registrados en el sistema.")
        st.markdown("</div>", unsafe_allow_html=True)

    elif menu_admin == "💸 Gastos Operativos":
        st.markdown(
            """
            <div class="top-banner">
                <div style="font-size: 18px; font-weight: 700;">Registrar Gastos u Operaciones</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        st.markdown('<div class="form-card">', unsafe_allow_html=True)
        cat_gasto = st.selectbox(
            "Categoría del Gasto",
            [
                "🏨 Hospedaje / Hotel",
                "🍽️ Comida / Alimentación",
                "🚗 Transporte / Combustible",
                "✈️ Fletes / Envíos USA-CR",
                "🛃 Aduana / Impuestos",
                "📦 Material de Empaque",
                "💡 Servicios y Operación",
                "🧩 Otros Gastos"
            ]
        )
        
        concepto = st.text_input("Concepto o Descripción", placeholder="Ej. Noche en hotel Miami, Almuerzo en gira de compras, Gasolina")
        monto_gasto = st.number_input("Monto en Colones (₡ CRC)", min_value=0.0, step=500.0)
        obs_gasto = st.text_area("Notas / Observaciones adicionales", height=60, placeholder="Ej. Factura #1024 / Pagado en efectivo")

        st.write("")
        if st.button("💾 Registrar Gasto Operativo"):
            if concepto and monto_gasto > 0:
                c.execute(
                    "INSERT INTO gastos (concepto, categoria, monto_crc, fecha, observaciones) VALUES (?, ?, ?, ?, ?)",
                    (concepto, cat_gasto, monto_gasto, datetime.now().strftime("%Y-%m-%d"), obs_gasto)
                )
                conn.commit()
                st.success("¡Gasto registrado e integrado a las finanzas correctamente!")
            else:
                st.error("Por favor completa la descripción y un monto superior a 0.")

        st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("<h4 style='color:#be185d;'>Últimos Gastos Registrados</h4>", unsafe_allow_html=True)
        gastos_df = pd.read_sql("SELECT fecha as Fecha, concepto as Concepto, categoria as Categoria, monto_crc as 'Monto (CRC)', observaciones as Observaciones FROM gastos ORDER BY id DESC LIMIT 10", conn)
        if not gastos_df.empty:
            st.dataframe(gastos_df, use_container_width=True)

    # --- CAMBIO 2: INTERFAZ DE EXPORTACIÓN LIMPIA ---
    elif menu_admin == "📊 Finanzas y Reportes":
        st.markdown(
            """
            <div class="top-banner">
                <div style="font-size: 18px; font-weight: 700;">Balance General y Ganancia Real</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        df_pedidos = pd.read_sql("SELECT COALESCE(SUM(precio * cantidad), 0) as total FROM productos", conn)
        df_abonos = pd.read_sql("SELECT COALESCE(SUM(monto_crc), 0) as total FROM abonos", conn)
        df_gastos = pd.read_sql("SELECT COALESCE(SUM(monto_crc), 0) as total FROM gastos", conn)

        total_ventas = df_pedidos.iloc[0]["total"]
        total_ingresos_reales = df_abonos.iloc[0]["total"]
        total_gastos = df_gastos.iloc[0]["total"]
        
        ganancia_real = total_ingresos_reales - total_gastos
        cuentas_por_cobrar = total_ventas - total_ingresos_reales

        col_f1, col_f2, col_f3 = st.columns(3)
        col_f1.metric("Ventas Totales", f"₡{total_ventas:,.0f}")
        col_f2.metric("Cobrado (Abonos)", f"₡{total_ingresos_reales:,.0f}")
        col_f3.metric("Por Cobrar", f"₡{max(0, cuentas_por_cobrar):,.0f}")

        col_f4, col_f5 = st.columns(2)
        col_f4.metric("Total Gastos (Egresos)", f"₡{total_gastos:,.0f}")
        col_f5.metric("💎 GANANCIA REAL NETO", f"₡{ganancia_real:,.0f}")

        st.write("")
        
        st.markdown('<div class="form-card">', unsafe_allow_html=True)
        st.markdown("<h4 style='color:#be185d;'>📊 Desglose de Gastos Operativos</h4>", unsafe_allow_html=True)
        df_gastos_cat = pd.read_sql("SELECT categoria as Categoría, SUM(monto_crc) as Total_CRC FROM gastos GROUP BY categoria ORDER BY Total_CRC DESC", conn)
        if not df_gastos_cat.empty:
            st.dataframe(df_gastos_cat, use_container_width=True)
        else:
            st.info("No hay gastos registrados en el sistema.")
        st.markdown("</div>", unsafe_allow_html=True)

        # Módulo de exportación con título limpio y sin el texto descriptivo
        st.markdown('<div class="form-card">', unsafe_allow_html=True)
        st.markdown("<h4 style='color:#be185d;'>📥 Exportar Reporte Financiero</h4>", unsafe_allow_html=True)

        output = io.BytesIO()
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill(start_color="DB2777", end_color="DB2777", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Segoe UI", size=10)
        title_font = Font(name="Segoe UI", size=14, bold=True, color="BE185D")
        thin_border = Border(
            left=Side(style='thin', color='FBCFE8'),
            right=Side(style='thin', color='FBCFE8'),
            top=Side(style='thin', color='FBCFE8'),
            bottom=Side(style='thin', color='FBCFE8')
        )

        tablas_config = [
            ("Resumen_Financiero", pd.DataFrame([
                {"Concepto": "Total Ventas Proyectadas", "Monto CRC": total_ventas},
                {"Concepto": "Ingresos Reales Cobrados (Abonos)", "Monto CRC": total_ingresos_reales},
                {"Concepto": "Cuentas Por Cobrar", "Monto CRC": max(0, cuentas_por_cobrar)},
                {"Concepto": "Total Gastos Operativos (Egresos)", "Monto CRC": total_gastos},
                {"Concepto": "GANANCIA REAL / UTILIDAD NETA", "Monto CRC": ganancia_real}
            ])),
            ("Clientes", pd.read_sql("SELECT id_cliente as Código, nombre as Nombre, telefono as Teléfono, correo as Correo FROM clientes", conn)),
            ("Ventas_Productos", pd.read_sql("SELECT p.id_cliente as Cliente, p.tienda as Tienda, p.categoria as Categoría, p.descripcion as Producto, p.precio as Precio_CRC, p.cantidad as Cantidad, (p.precio * p.cantidad) as Total_CRC, p.estado as Estado, p.id_caja as Caja FROM productos p", conn)),
            ("Ingresos_Abonos", pd.read_sql("SELECT id_cliente as Cliente, monto_crc as Monto_CRC, fecha as Fecha FROM abonos", conn)),
            ("Gastos_Operativos", pd.read_sql("SELECT fecha as Fecha, categoria as Categoría, concepto as Concepto, monto_crc as Monto_CRC, observaciones as Observaciones FROM gastos", conn))
        ]

        for sheet_name, df_data in tablas_config:
            ws = wb.create_sheet(title=sheet_name)
            ws.views.sheetView[0].showGridLines = True
            
            ws.cell(row=1, column=1, value=f"Reporte: {sheet_name.replace('_', ' ')}").font = title_font
            
            if not df_data.empty:
                headers = list(df_data.columns)
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                for row_idx, row_data in enumerate(df_data.values, 4):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.font = data_font
                        cell.border = thin_border
                        
                        if "CRC" in headers[col_idx-1] or "Monto" in headers[col_idx-1] or "Precio" in headers[col_idx-1] or "Total" in headers[col_idx-1]:
                            cell.number_format = '₡#,##0'
                            cell.alignment = Alignment(horizontal="right")

                for col in ws.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(output)
        excel_data = output.getvalue()

        st.download_button(
            label="📊 Descargar Reporte (.xlsx)",
            data=excel_data,
            file_name=f"Reporte_Financiero_Minici_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.markdown("</div>", unsafe_allow_html=True)

# --- PANEL CLIENTA ---
elif st.session_state.user_role == "client":
    id_cli = st.session_state.current_client

    c.execute("SELECT nombre FROM clientes WHERE id_cliente = ?", (id_cli,))
    res_cli = c.fetchone()
    client_name = res_cli[0] if res_cli else "Clienta"

    prods = pd.read_sql(
        "SELECT descripcion, tienda, precio, estado, foto_path FROM productos WHERE id_cliente = ?",
        conn,
        params=(id_cli,),
    )
    abonos_df = pd.read_sql(
        "SELECT COALESCE(SUM(monto_crc), 0.0) as total FROM abonos WHERE id_cliente = ?",
        conn,
        params=(id_cli,),
    )
    total_abonos = abonos_df.iloc[0]["total"]
    total_compras_crc = prods["precio"].sum() if not prods.empty else 0.0

    st.markdown(
        f"""
    <div class="top-banner">
        <div style="font-size: 20px; font-weight: bold;">Hola, {client_name} 👋</div>
        <div style="font-size: 13px;">Tu código: <b>{id_cli}</b></div>
    </div>
    """,
        unsafe_allow_html=True,
    )

    if st.button("🚪 Salir de mi cuenta"):
        st.session_state.user_role = None
        st.session_state.current_client = None
        st.rerun()

    st.markdown("<h3 style='color:#db2777;'>📦 Mis Pedidos</h3>", unsafe_allow_html=True)

    col_m1, col_m2, col_m3 = st.columns(3)
    col_m1.metric("Total Pedidos", f"₡{total_compras_crc:,.0f}")
    col_m2.metric("Total Abonado", f"₡{total_abonos:,.0f}")
    col_m3.metric("Saldo Pendiente", f"₡{max(0, total_compras_crc - total_abonos):,.0f}")

    if prods.empty:
        st.info("Aún no tienes productos registrados.")
    else:
        for idx, row in prods.iterrows():
            precio_val = row['precio'] if pd.notnull(row['precio']) else 0.0
            st.markdown(
                f"""
            <div class="form-card" style="margin-bottom: 10px; padding: 12px;">
                <div style="font-weight: bold; font-size: 15px; color:#0f172a;">{row['descripcion']}</div>
                <div style="font-size: 12px; color: #64748b; margin-top: 2px;">Tienda: <b>{row['tienda']}</b> | Precio: ₡{precio_val:,.0f}</div>
                <div style="margin-top: 6px;"><span style="background: #fdf2f8; color: #db2777; padding: 2px 6px; border-radius: 6px; font-size: 11px; font-weight: bold;">{row['estado']}</span></div>
            </div>
            """,
                unsafe_allow_html=True,
            )
            if row["foto_path"] and os.path.exists(
                os.path.join("fotos_productos", row["foto_path"])
            ):
                st.image(os.path.join("fotos_productos", row["foto_path"]), width=120)